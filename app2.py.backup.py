"""
Analizador Comparativo de Pólizas + OCR
Backend Flask con toda la lógica de procesamiento
"""
import os
import fitz  # PyMuPDF
import numpy as np
from sklearn.metrics.pairwise import cosine_similarity
from mistralai import Mistral
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
import time
import random
import re
import tempfile
import pandas as pd
import io
from PIL import Image
import pytesseract
from flask import Flask, render_template, request, send_file, flash, redirect, url_for
from werkzeug.utils import secure_filename
import secrets

# =========================
# CONFIGURACIÓN FLASK
# =========================
app = Flask(__name__)
app.secret_key = secrets.token_hex(16)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100 MB max file size
app.config['UPLOAD_FOLDER'] = tempfile.mkdtemp()

# =========================
# CONFIGURACIÓN DE PROCESAMIENTO
# =========================
EMBED_MODEL = "mistral-embed"
LLM_MODEL = "mistral-small-latest"

# Configuración de columnas Excel
FILA_INICIO = 17
COL_A = 1
COL_B = 2
COL_G = 7
COL_H = 8
COL_I = 9
COL_F = 6

PAUSA_ENTRE_CONSULTAS = 0.5
MAX_REINTENTOS = 2

# Cache para embeddings
embeddings_cache_pdf1 = {}
embeddings_cache_pdf2 = {}
embeddings_cache_ocr = {}

# =========================
# FUNCIÓN DE EXTRACCIÓN OCR INTELIGENTE
# =========================
def extraer_texto_pdf(path):
    """Extrae texto con detección automática de PDFs escaneados"""
    doc = fitz.open(path)

    # Detectar si es PDF escaneado
    tiene_texto_nativo = False
    for i in range(min(3, len(doc))):
        if doc[i].get_text().strip():
            tiene_texto_nativo = True
            break

    if tiene_texto_nativo:
        # Extracción nativa rápida
        texto = ""
        for page in doc:
            texto += page.get_text() + "\n\n"
        doc.close()
        return texto.strip()

    # OCR solo si es necesario
    print(f"🖼️ PDF detectado como escaneado. Aplicando OCR...")
    texto = ""
    for page_num in range(len(doc)):
        pix = doc[page_num].get_pixmap(dpi=200)
        img = Image.open(io.BytesIO(pix.tobytes()))
        texto_pagina = pytesseract.image_to_string(
            img,
            lang="spa+eng",
            config="--oem 1 --psm 6"
        )
        texto += f"\n--- PÁGINA {page_num + 1} ---\n" + texto_pagina + "\n\n"
    doc.close()
    return texto.strip()

# =========================
# INICIALIZACIÓN MISTRAL
# =========================
def inicializar_mistral():
    """Inicializa el cliente Mistral con API key de secrets"""
    try:
        from dotenv import load_dotenv
        load_dotenv()
        api_key = os.getenv("MISTRAL_API_KEY")

        if not api_key:
            # Intentar desde archivo secrets.toml
            secrets_path = os.path.join(os.path.dirname(__file__), '.streamlit', 'secrets.toml')
            if os.path.exists(secrets_path):
                import toml
                secrets_data = toml.load(secrets_path)
                api_key = secrets_data.get('MISTRAL_API_KEY', '')

        if not api_key:
            raise Exception("API Key no encontrada")

        client = Mistral(api_key=api_key)
        # Verificar conexión
        client.models.list()
        return client
    except Exception as e:
        print(f"Error al inicializar Mistral: {e}")
        return None

# =========================
# FUNCIONES PARA EXCEL
# =========================
def valor_real(ws, celda):
    """Obtiene el valor real de una celda, manejando celdas fusionadas"""
    if not isinstance(celda, MergedCell):
        return celda.value
    for r in ws.merged_cells.ranges:
        if celda.coordinate in r:
            return ws.cell(r.min_row, r.min_col).value
    return None

def construir_vector(ws):
    """Construye un vector con las filas y textos a procesar del Excel"""
    vector = []
    for row in range(FILA_INICIO, ws.max_row + 1):
        a = valor_real(ws, ws.cell(row=row, column=COL_A))
        b = ws.cell(row=row, column=COL_B).value

        if a and str(a).strip():
            texto = f"{a} {b}".strip() if b else str(a)
            vector.append((row, texto))

    return vector

def escribir_en_celda(ws, row, col, valor):
    """Escribe un valor en una celda, manejando celdas fusionadas"""
    celda = ws.cell(row=row, column=col)
    if not isinstance(celda, MergedCell):
        celda.value = valor
        return
    for r in ws.merged_cells.ranges:
        if celda.coordinate in r:
            ws.cell(r.min_row, r.min_col).value = valor
            return

# =========================
# FUNCIONES PARA PDF (RAG)
# =========================
def dividir_texto(texto, chunk_size=1800, overlap=250):
    """Divide el texto en chunks para procesamiento"""
    chunks = []
    start = 0
    while start < len(texto):
        end = start + chunk_size
        chunks.append(texto[start:end])
        start = end - overlap
    return chunks

def generar_embeddings_con_reintentos(textos, client):
    """Genera embeddings con reintentos optimizados"""
    if isinstance(textos, str):
        textos = [textos]

    for intento in range(MAX_REINTENTOS + 1):
        try:
            response = client.embeddings.create(
                model=EMBED_MODEL,
                inputs=textos
            )
            return np.array([e.embedding for e in response.data])
        except Exception as e:
            if intento < MAX_REINTENTOS and ("rate_limited" in str(e) or "429" in str(e)):
                espera = (2 ** intento) * 0.5 + random.uniform(0.5, 1.5)
                time.sleep(espera)
            else:
                raise e

    raise Exception(f"Error después de {MAX_REINTENTOS} reintentos")

def buscar_contexto(pregunta, chunks, embeddings, embeddings_cache, client, top_k=5, umbral=0.15):
    """Busca contexto relevante optimizado"""
    if pregunta not in embeddings_cache:
        embeddings_cache[pregunta] = generar_embeddings_con_reintentos([pregunta], client)[0]

    emb_q = embeddings_cache[pregunta].reshape(1, -1)
    sims = cosine_similarity(emb_q, embeddings)[0]
    idx = sims.argsort()[::-1]

    seleccion = []
    for i in idx[:top_k*2]:
        if sims[i] >= umbral or len(seleccion) < 2:
            seleccion.append(chunks[i])
        if len(seleccion) == top_k:
            break

    return seleccion

# =========================
# DICCIONARIO DE TÉRMINOS DE SEGUROS
# =========================
TERMINOS_SEGUROS = {
    "modalidad": ["modalidad", "tipo de póliza", "sistema", "modelo", "forma", "tipo", "clase", "categoría"],
    "mixto": ["mixto", "combinado", "híbrido", "mixta"],
    "abierto": ["sistema abierto", "abierto", "libre elección", "elección libre"],
    "cerrado": ["sistema cerrado", "cerrado", "red cerrada", "proveedores específicos"],
    "alcance de la cobertura": ["cobertura geográfica", "ámbito geográfico", "alcance territorial",
                               "cobertura territorial", "ámbito de cobertura", "zona de cobertura"],
    "nacional": ["nacional", "todo el país", "en todo bolivia", "a nivel nacional"],
    "internacional": ["internacional", "en el extranjero", "fuera del país", "cobertura internacional"],
    "departamentos": ["departamentos", "regiones", "provincias", "ciudades"],
    "cobertura": ["cobertura", "protección", "amparo", "beneficio", "garantía", "inclusión"],
    "capital": ["capital asegurado", "monto asegurado", "suma asegurada", "valor asegurado",
               "límite", "importe", "cantidad", "capital"],
    "prima": ["prima", "precio", "costo", "tarifa", "valor", "cuota"],
    "deducible": ["deducible", "franquicia", "copago", "participación"],
    "reembolso": ["reembolso", "devolución", "pago", "compensación", "restitución"],
    "hospitalización": ["hospitalización", "internación", "ingreso", "estancia hospitalaria"],
    "ambulatorio": ["ambulatorio", "consulta externa", "tratamiento ambulatorio"],
    "emergencia": ["emergencia", "urgencia", "accidente", "siniestro"],
    "exclusión": ["exclusión", "no cubre", "excluye", "no incluye", "excepto", "limitación"],
    "covid": ["covid", "coronavirus", "pandemia", "enfermedad epidémica"],
    "maternidad": ["maternidad", "embarazo", "parto", "nacimiento", "control prenatal"],
    "parto": ["parto natural", "parto normal", "parto", "nacimiento"],
}

# =========================
# FUNCIONES DE ANÁLISIS LLM
# =========================
def detectar_tipo_consulta(pregunta):
    """Detecta el tipo de consulta para usar sinónimos específicos"""
    pregunta_lower = pregunta.lower()
    tipos_detectados = []

    for termino, sinonimos in TERMINOS_SEGUROS.items():
        if any(s in pregunta_lower for s in [termino] + sinonimos[:2]):
            tipos_detectados.append(termino)

    if "modalidad" in tipos_detectados:
        return "MODALIDAD"
    elif "geográfico" in tipos_detectados or "nacional" in tipos_detectados or "internacional" in tipos_detectados:
        return "COBERTURA_GEOGRÁFICA"
    elif "capital" in tipos_detectados:
        return "MONTO_CAPITAL"
    elif "reembolso" in tipos_detectados:
        return "REEMBOLSO"
    elif "cobertura" in tipos_detectados:
        return "COBERTURA_GENERAL"

    return "GENERAL"

def preguntar_llm_optimizada(pregunta, contexto, client):
    """LLM optimizado con diccionario mejorado de términos de seguros"""
    tipo_consulta = detectar_tipo_consulta(pregunta)

    sinonimos_relevantes = []
    for termino, sinonimos in TERMINOS_SEGUROS.items():
        if termino in pregunta.lower() or any(s in pregunta.lower() for s in sinonimos[:3]):
            sinonimos_relevantes.extend(sinonimos[:3])

    sinonimos_texto = ", ".join(set(sinonimos_relevantes[:6]))

    prompt = f"""
Eres experto en pólizas de seguros. Analiza la siguiente consulta y busca información que coincida o similitudes.

DOCUMENTO DE LA PÓLIZA:
{contexto}

CONSULTA: "{pregunta}"

TIPO DE CONSULTA DETECTADA: {tipo_consulta}
TÉRMINOS RELACIONADOS PARA BÚSQUEDA: {sinonimos_texto}

INSTRUCCIONES ESPECÍFICAS PARA {tipo_consulta}:

{get_instrucciones_especificas(tipo_consulta)}

REGLAS GENERALES:
1. Busca por SIGNIFICADO, no solo por palabras exactas
2. Usa los términos relacionados para búsqueda ampliada
3. Copia EXACTAMENTE el texto relevante del documento
4. NO inventes, NO interpretes, NO resumas
5. Si no encuentras, responde EXACTAMENTE: "no encontrado"
6. NO agregues "Respuesta:", ni explicaciones, ni formato

RESPUESTA DIRECTA DEL DOCUMENTO:
"""

    for intento in range(MAX_REINTENTOS + 1):
        try:
            response = client.chat.complete(
                model=LLM_MODEL,
                messages=[{"role": "user", "content": prompt}],
                temperature=0.1,
                max_tokens=150
            )

            respuesta = response.choices[0].message.content.strip()
            return clean_response(respuesta)

        except Exception as e:
            if intento < MAX_REINTENTOS and ("rate_limited" in str(e) or "429" in str(e)):
                espera = (2 ** intento) * 0.3 + random.uniform(0.3, 1.0)
                time.sleep(espera)
            else:
                return "error"

    return "error: rate limit"

def get_instrucciones_especificas(tipo_consulta):
    """Devuelve instrucciones específicas para cada tipo de consulta"""
    instrucciones = {
        "MODALIDAD": """
        • Busca términos: "modalidad", "tipo de póliza", "sistema"
        • Especifica si es: mixto, abierto, cerrado, combinado
        • Incluye detalles del sistema de atención
        """,
        "COBERTURA_GEOGRÁFICA": """
        • Busca términos: "cobertura geográfica", "ámbito", "departamentos", "nacional", "internacional"
        • Especifica zonas, regiones, países cubiertos
        • Menciona límites territoriales si existen
        """,
        "MONTO_CAPITAL": """
        • Busca términos: "capital", "monto", "suma", "valor", "USD", "Bs.", "límite"
        • Extrae números exactos con su moneda
        • Incluye condiciones si las hay
        """,
        "REEMBOLSO": """
        • Busca términos: "reembolso", "porcentaje", "%", "cobertura", "pago"
        • Extrae porcentajes exactos y condiciones
        • Especifica plazos y modalidades
        """,
        "COBERTURA_GENERAL": """
        • Busca términos específicos de la consulta
        • Extrae condiciones, límites, inclusiones
        • Especifica detalles relevantes
        """,
        "GENERAL": """
        • Busca información relacionada con la consulta
        • Usa términos equivalentes del sector
        • Extrae información precisa y relevante
        """
    }

    return instrucciones.get(tipo_consulta, instrucciones["GENERAL"])

def clean_response(respuesta):
    """Limpia la respuesta de forma robusta"""
    if not respuesta:
        return "no encontrado"

    respuesta_lower = respuesta.lower()

    no_encontrado_patterns = [
        "no encontrado", "no se encuentra", "no existe", "no hay información",
        "no aparece", "no figura", "no se menciona", "no se halla",
        "información no disponible", "no disponible"
    ]

    for pattern in no_encontrado_patterns:
        if pattern in respuesta_lower:
            return "no encontrado"

    prefixes_to_remove = [
        r'^respuesta[:\s]*',
        r'^resultado[:\s]*',
        r'^encontrado[:\s]*',
        r'^información[:\s]*',
        r'^\s*[-*•]\s*',
        r'^\s*\d+[\.\)]\s*'
    ]

    for prefix in prefixes_to_remove:
        respuesta = re.sub(prefix, '', respuesta, flags=re.IGNORECASE)

    respuesta = re.sub(r'\*\*.*?\*\*', '', respuesta)
    respuesta = re.sub(r'__.*?__', '', respuesta)
    respuesta = re.sub(r'\s+', ' ', respuesta).strip()

    if len(respuesta) < 10 or len(respuesta.split()) < 3:
        return "no encontrado"

    if len(respuesta) > 300:
        sentences = re.split(r'[.!?]', respuesta)
        for sentence in sentences:
            if re.search(r'[\d\.,]+', sentence) and len(sentence) > 20:
                respuesta = sentence.strip() + "."
                break

        if len(respuesta) > 300:
            respuesta = respuesta[:297] + "..."

    return respuesta

# =========================
# FUNCIONES DE PREPARACIÓN
# =========================
def preparar_documento(pdf_path, client):
    """Prepara un documento PDF para búsqueda RAG"""
    texto = extraer_texto_pdf(pdf_path)
    chunks = dividir_texto(texto, chunk_size=2000, overlap=300)

    embeddings_list = []
    batch_size = 8

    for i in range(0, len(chunks), batch_size):
        batch = chunks[i:i+batch_size]
        try:
            batch_embeddings = generar_embeddings_con_reintentos(batch, client)
            embeddings_list.append(batch_embeddings)
            time.sleep(0.3)
        except Exception:
            continue

    embeddings = np.vstack(embeddings_list) if embeddings_list else np.array([])
    return chunks, embeddings

def procesar_consulta_dual(fila, consulta, chunks1, embeddings1, chunks2, embeddings2, chunks_ocr, embeddings_ocr, client, usar_ocr):
    """Procesa una consulta en ambos PDFs + opcional OCR"""
    resultado = {'fila': fila, 'consulta': consulta}

    # Procesar con PDF 1
    try:
        contexto_chunks1 = buscar_contexto(consulta, chunks1, embeddings1, embeddings_cache_pdf1, client)
        if contexto_chunks1:
            contexto_texto1 = "\n\n".join(contexto_chunks1[:3])
            respuesta1 = preguntar_llm_optimizada(consulta, contexto_texto1, client)
        else:
            respuesta1 = "no encontrado"
    except Exception as e:
        respuesta1 = f"error: {str(e)[:50]}"

    # Procesar con PDF 2
    try:
        contexto_chunks2 = buscar_contexto(consulta, chunks2, embeddings2, embeddings_cache_pdf2, client)
        if contexto_chunks2:
            contexto_texto2 = "\n\n".join(contexto_chunks2[:3])
            respuesta2 = preguntar_llm_optimizada(consulta, contexto_texto2, client)
        else:
            respuesta2 = "no encontrado"
    except Exception as e:
        respuesta2 = f"error: {str(e)[:50]}"

    # Procesar con OCR si está habilitado
    if usar_ocr and chunks_ocr is not None and len(embeddings_ocr) > 0:
        try:
            contexto_chunks3 = buscar_contexto(consulta, chunks_ocr, embeddings_ocr, embeddings_cache_ocr, client)
            if contexto_chunks3:
                contexto_texto3 = "\n\n".join(contexto_chunks3[:3])
                respuesta3 = preguntar_llm_optimizada(consulta, contexto_texto3, client)
            else:
                respuesta3 = "no encontrado"
        except Exception as e:
            respuesta3 = f"error: {str(e)[:50]}"
    else:
        respuesta3 = None

    resultado['respuesta_pdf1'] = respuesta1
    resultado['respuesta_pdf2'] = respuesta2
    resultado['respuesta_ocr'] = respuesta3

    return resultado

# =========================
# RUTAS FLASK
# =========================

@app.route('/')
def index():
    """Página principal con formulario"""
    return render_template('index.html')

@app.route('/procesar', methods=['POST'])
def procesar():
    """Procesa los archivos subidos y muestra resultados"""
    try:
        # Verificar que se hayan subido los archivos requeridos
        if 'pdf1' not in request.files or 'pdf2' not in request.files or 'excel' not in request.files:
            flash('❌ Por favor, sube todos los archivos requeridos.', 'error')
            return redirect(url_for('index'))

        pdf1 = request.files['pdf1']
        pdf2 = request.files['pdf2']
        excel = request.files['excel']
        pdf3 = request.files.get('pdf3')  # Opcional

        if pdf1.filename == '' or pdf2.filename == '' or excel.filename == '':
            flash('❌ Por favor, selecciona todos los archivos requeridos.', 'error')
            return redirect(url_for('index'))

        # Crear directorio temporal
        tmpdir = tempfile.mkdtemp()

        # Guardar archivos
        pdf1_path = os.path.join(tmpdir, secure_filename(pdf1.filename))
        pdf2_path = os.path.join(tmpdir, secure_filename(pdf2.filename))
        excel_path = os.path.join(tmpdir, secure_filename(excel.filename))

        pdf1.save(pdf1_path)
        pdf2.save(pdf2_path)
        excel.save(excel_path)

        pdf3_path = None
        usar_ocr = False
        if pdf3 and pdf3.filename != '':
            pdf3_path = os.path.join(tmpdir, secure_filename(pdf3.filename))
            pdf3.save(pdf3_path)
            usar_ocr = True

        # Inicializar cliente Mistral
        client = inicializar_mistral()
        if not client:
            flash('❌ Error al conectar con Mistral AI. Verifica tu API Key.', 'error')
            return redirect(url_for('index'))

        # 1. Cargar Excel
        wb = load_workbook(excel_path)
        ws = wb.active
        vector = construir_vector(ws)
        total_filas = len(vector)

        if total_filas == 0:
            flash('❌ No se encontraron filas para procesar en el Excel.', 'error')
            return redirect(url_for('index'))

        # 2. Preparar PDF 1
        chunks1, embeddings1 = preparar_documento(pdf1_path, client)

        # 3. Preparar PDF 2
        chunks2, embeddings2 = preparar_documento(pdf2_path, client)

        # 4. Preparar PDF OCR si existe
        chunks_ocr, embeddings_ocr = None, np.array([])
        if usar_ocr:
            chunks_ocr, embeddings_ocr = preparar_documento(pdf3_path, client)

        # 5. Procesar filas
        resultados = []
        inicio_proc = time.time()

        for i, (fila, texto_consulta) in enumerate(vector, 1):
            if i > 1:
                time.sleep(PAUSA_ENTRE_CONSULTAS)

            resultado = procesar_consulta_dual(
                fila, texto_consulta,
                chunks1, embeddings1,
                chunks2, embeddings2,
                chunks_ocr, embeddings_ocr,
                client,
                usar_ocr
            )
            resultados.append(resultado)

        tiempo_proc = time.time() - inicio_proc

        # 6. Guardar resultados
        ws.cell(row=FILA_INICIO-1, column=COL_F, value="DOCUMENTO")
        ws.cell(row=FILA_INICIO-1, column=COL_G, value=pdf1.filename.replace(".pdf", ""))
        ws.cell(row=FILA_INICIO-1, column=COL_H, value=pdf2.filename.replace(".pdf", ""))

        if usar_ocr:
            ws.cell(row=FILA_INICIO-1, column=COL_I, value="OCR: " + pdf3.filename.replace(".pdf", ""))

        for r in resultados:
            escribir_en_celda(ws, r['fila'], COL_G, r['respuesta_pdf1'])
            escribir_en_celda(ws, r['fila'], COL_H, r['respuesta_pdf2'])
            if usar_ocr and r['respuesta_ocr']:
                escribir_en_celda(ws, r['fila'], COL_I, r['respuesta_ocr'])

        output_path = os.path.join(tmpdir, "resultados_comparacion.xlsx")
        wb.save(output_path)

        # 7. Calcular estadísticas
        contadores_pdf1 = {'encontrado': 0, 'no_encontrado': 0, 'error': 0}
        contadores_pdf2 = {'encontrado': 0, 'no_encontrado': 0, 'error': 0}
        contadores_ocr = {'encontrado': 0, 'no_encontrado': 0, 'error': 0} if usar_ocr else None

        for r in resultados:
            if "no encontrado" in str(r['respuesta_pdf1']).lower():
                contadores_pdf1['no_encontrado'] += 1
            elif "error" in str(r['respuesta_pdf1']).lower():
                contadores_pdf1['error'] += 1
            else:
                contadores_pdf1['encontrado'] += 1

            if "no encontrado" in str(r['respuesta_pdf2']).lower():
                contadores_pdf2['no_encontrado'] += 1
            elif "error" in str(r['respuesta_pdf2']).lower():
                contadores_pdf2['error'] += 1
            else:
                contadores_pdf2['encontrado'] += 1

            if usar_ocr and r['respuesta_ocr']:
                if "no encontrado" in str(r['respuesta_ocr']).lower():
                    contadores_ocr['no_encontrado'] += 1
                elif "error" in str(r['respuesta_ocr']).lower():
                    contadores_ocr['error'] += 1
                else:
                    contadores_ocr['encontrado'] += 1

        # 8. Preparar datos para la plantilla
        velocidad = (total_filas / tiempo_proc) * 60 if tiempo_proc > 0 else 0

        # Guardar el archivo de resultados con un nombre único
        result_filename = f"resultados_{int(time.time())}.xlsx"
        result_path = os.path.join(app.config['UPLOAD_FOLDER'], result_filename)
        os.rename(output_path, result_path)

        # Limpiar caches
        embeddings_cache_pdf1.clear()
        embeddings_cache_pdf2.clear()
        embeddings_cache_ocr.clear()

        return render_template('results.html',
            tiempo_min=tiempo_proc / 60,
            velocidad=velocidad,
            total_filas=total_filas,
            pdf1_name=pdf1.filename,
            pdf2_name=pdf2.filename,
            pdf3_name=pdf3.filename if usar_ocr else None,
            contadores_pdf1=contadores_pdf1,
            contadores_pdf2=contadores_pdf2,
            contadores_ocr=contadores_ocr,
            resultados=resultados[:10],  # Solo primeras 10 filas
            result_filename=result_filename
        )

    except Exception as e:
        flash(f'❌ Error durante el procesamiento: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/descargar/<filename>')
def descargar(filename):
    """Descarga el archivo de resultados"""
    try:
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        if os.path.exists(file_path):
            return send_file(file_path, as_attachment=True)
        else:
            flash('❌ Archivo no encontrado.', 'error')
            return redirect(url_for('index'))
    except Exception as e:
        flash(f'❌ Error al descargar: {str(e)}', 'error')
        return redirect(url_for('index'))

# =========================
# EJECUCIÓN
# =========================
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8501, debug=True)
