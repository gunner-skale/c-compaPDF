"""
Analizador Comparativo de Pólizas + OCR
Backend Flask optimizado para velocidad máxima
"""
import os
import fitz  # PyMuPDF
import numpy as np
from sklearn.metrics.pairwise import cosine_similarity
from mistralai import Mistral
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
import time
import random
import re
import tempfile
import pandas as pd
import io
from PIL import Image
import pytesseract
from flask import Flask, render_template, request, send_file, jsonify, redirect, url_for, flash
from werkzeug.utils import secure_filename
import threading
import json
from datetime import datetime

# =========================
# CONFIGURACIÓN FLASK
# =========================
app = Flask(__name__)
app.secret_key = os.urandom(24)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100 MB
app.config['UPLOAD_FOLDER'] = tempfile.mkdtemp()
app.config['TASKS'] = {}  # {task_id: {status, progress, message, result, error}}

# =========================
# CONFIGURACIÓN OPTIMIZADA PARA VELOCIDAD
# =========================
EMBED_MODEL = "mistral-embed"
LLM_MODEL = "mistral-small-latest"  # ✅ CORREGIDO: open-mixtral-8x7b no existe (causaba todos los errores)

# Configuración de columnas Excel
FILA_INICIO = 17
COL_A = 1
COL_B = 2
COL_G = 7
COL_H = 8
COL_I = 9
COL_F = 6

# ⚡ PARÁMETROS OPTIMIZADOS PARA VELOCIDAD
PAUSA_ENTRE_CONSULTAS = 0.2
MAX_REINTENTOS = 2
CHUNK_SIZE = 2000
OVERLAP =200
TOP_K = 5
UMBRAL_SIMILITUD = 0.25 #0.15
BATCH_SIZE_EMBEDDINGS = 12

# Caches para embeddings
embeddings_cache_pdf1 = {}
embeddings_cache_pdf2 = {}
embeddings_cache_ocr = {}

# =========================
# FUNCIÓN DE EXTRACCIÓN OCR INTELIGENTE
# =========================
def extraer_texto_pdf(path, task_id=None, stage=""):
    """Extrae texto con detección automática + actualiza progreso"""
    if task_id:
        update_task(task_id, progress=5, message=f"{stage}Analizando PDF...")

    doc = fitz.open(path)

    # Detectar si es PDF escaneado (sin texto en primeras 3 páginas)
    tiene_texto_nativo = False
    for i in range(min(3, len(doc))):
        if doc[i].get_text().strip():
            tiene_texto_nativo = True
            break

    if tiene_texto_nativo:
        # Extracción nativa RÁPIDA
        texto = ""
        total_pages = len(doc)
        for page_num, page in enumerate(doc):
            if task_id:
                progress_base = 5 if "Póliza 1" in stage else 25
                progress = progress_base + (page_num / total_pages) * 15
                update_task(task_id, progress=progress, message=f"{stage}Página {page_num+1}/{total_pages}")
            texto += page.get_text() + "\n\n"
        doc.close()
        return texto.strip()

    # OCR SOLO si es necesario
    if task_id:
        update_task(task_id, message=f"{stage}🖼️ PDF escaneado detectado. Aplicando OCR...")

    texto = ""
    total_pages = len(doc)
    for page_num in range(total_pages):
        if task_id:
            progress_base = 5 if "Póliza 1" in stage else 25
            progress = progress_base + (page_num / total_pages) * 15
            update_task(task_id, progress=progress, message=f"{stage}OCR página {page_num+1}/{total_pages}")

        pix = doc[page_num].get_pixmap(dpi=200)
        img = Image.open(io.BytesIO(pix.tobytes()))
        texto_pagina = pytesseract.image_to_string(
            img,
            lang="spa+eng",
            config="--oem 1 --psm 6"
        )
        texto += f"\n--- PÁGINA {page_num + 1} ---\n" + texto_pagina + "\n\n"
    doc.close()
    return texto.strip()

def update_task(task_id, **kwargs):
    """Actualiza el estado de una tarea en background"""
    if task_id in app.config['TASKS']:
        for key, value in kwargs.items():
            app.config['TASKS'][task_id][key] = value

# =========================
# INICIALIZACIÓN MISTRAL
# =========================
def inicializar_mistral():
    """Inicializa el cliente Mistral con API key de .env"""
    api_key = os.getenv("MISTRAL_API_KEY")
    if not api_key:
        raise Exception("MISTRAL_API_KEY no encontrada en variables de entorno")

    client = Mistral(api_key=api_key)
    client.models.list()  # Verificar conexión
    return client

# =========================
# FUNCIONES PARA EXCEL
# =========================
def valor_real(ws, celda):
    if not isinstance(celda, MergedCell):
        return celda.value
    for r in ws.merged_cells.ranges:
        if celda.coordinate in r:
            return ws.cell(r.min_row, r.min_col).value
    return None

def construir_vector(ws):
    vector = []
    for row in range(FILA_INICIO, ws.max_row + 1):
        a = valor_real(ws, ws.cell(row=row, column=COL_A))
        b = ws.cell(row=row, column=COL_B).value
        if a and str(a).strip():
            texto = f"{a} {b}".strip() if b else str(a)
            vector.append((row, texto))
    return vector

def escribir_en_celda(ws, row, col, valor):
    celda = ws.cell(row=row, column=col)
    if not isinstance(celda, MergedCell):
        celda.value = valor
        return
    for r in ws.merged_cells.ranges:
        if celda.coordinate in r:
            ws.cell(r.min_row, r.min_col).value = valor
            return

# ✅ NUEVA FUNCIÓN: DUPLICAR HOJA EXACTAMENTE (incluyendo estilos, fusiones y formato)
def duplicar_hoja(wb, hoja_origen, nombre_nuevo):
    """Crea una copia EXACTA de una hoja (incluyendo estilos, fusiones y formato)"""
    hoja_nueva = wb.create_sheet(title=nombre_nuevo)

    # Copiar todas las celdas con sus valores y estilos
    for row in hoja_origen.iter_rows(min_row=1, max_row=hoja_origen.max_row, min_col=1, max_col=hoja_origen.max_column):
        for cell in row:
            nueva_celda = hoja_nueva.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                nueva_celda.font = cell.font.copy()
                nueva_celda.border = cell.border.copy()
                nueva_celda.fill = cell.fill.copy()
                nueva_celda.number_format = cell.number_format
                nueva_celda.protection = cell.protection.copy()
                nueva_celda.alignment = cell.alignment.copy()

    # Copiar celdas fusionadas
    for rango in hoja_origen.merged_cells.ranges:
        hoja_nueva.merge_cells(str(rango))

    # Copiar dimensiones de columnas
    for col_letter in hoja_origen.column_dimensions:
        if col_letter in hoja_origen.column_dimensions:
            hoja_nueva.column_dimensions[col_letter] = hoja_origen.column_dimensions[col_letter]

    # Copiar dimensiones de filas
    for row_number in hoja_origen.row_dimensions:
        if row_number in hoja_origen.row_dimensions:
            hoja_nueva.row_dimensions[row_number] = hoja_origen.row_dimensions[row_number]

    return hoja_nueva

# =========================
# FUNCIONES PARA PDF (RAG) - OPTIMIZADAS
# =========================
def dividir_texto(texto, chunk_size=CHUNK_SIZE, overlap=OVERLAP):
    chunks = []
    start = 0
    while start < len(texto):
        end = start + chunk_size
        chunks.append(texto[start:end])
        start = end - overlap
    return chunks

def generar_embeddings_con_reintentos(textos, client):
    if isinstance(textos, str):
        textos = [textos]

    for intento in range(MAX_REINTENTOS + 1):
        try:
            response = client.embeddings.create(model=EMBED_MODEL, inputs=textos)
            return np.array([e.embedding for e in response.data])
        except Exception as e:
            if intento < MAX_REINTENTOS and ("rate_limited" in str(e) or "429" in str(e)):
                espera = (2 ** intento) * 0.3 + random.uniform(0.3, 0.8)
                time.sleep(espera)
            else:
                raise e

    raise Exception(f"Error después de {MAX_REINTENTOS} reintentos")

def buscar_contexto(pregunta, chunks, embeddings, embeddings_cache, client, top_k=TOP_K, umbral=UMBRAL_SIMILITUD):
    if pregunta not in embeddings_cache:
        embeddings_cache[pregunta] = generar_embeddings_con_reintentos([pregunta], client)[0]

    emb_q = embeddings_cache[pregunta].reshape(1, -1)
    sims = cosine_similarity(emb_q, embeddings)[0]
    idx = sims.argsort()[::-1]

    seleccion = []
    for i in idx[:top_k*3]:
        if sims[i] >= umbral or len(seleccion) < 3:
            seleccion.append(chunks[i])
        if len(seleccion) == top_k:
            break

    return seleccion

# =========================
# DICCIONARIO AMPLIADO DE TÉRMINOS DE SEGUROS
# =========================
TERMINOS_SEGUROS = {
    "modalidad": ["modalidad", "tipo de póliza", "sistema", "modelo", "forma", "tipo", "clase", "categoría", "plan"],
    "mixto": ["mixto", "combinado", "híbrido", "mixta", "dual"],
    "abierto": ["sistema abierto", "abierto", "libre elección", "elección libre", "red abierta"],
    "cerrado": ["sistema cerrado", "cerrado", "red cerrada", "proveedores específicos", "lista restringida"],
    "alcance de la cobertura": ["cobertura geográfica", "ámbito geográfico", "alcance territorial", "cobertura territorial", "ámbito de cobertura", "zona de cobertura", "territorio"],
    "nacional": ["nacional", "todo el país", "en todo bolivia", "a nivel nacional", "territorio nacional"],
    "internacional": ["internacional", "en el extranjero", "fuera del país", "cobertura internacional", "global"],
    "departamentos": ["departamentos", "regiones", "provincias", "ciudades", "estados", "municipios"],
    "cobertura": ["cobertura", "protección", "amparo", "beneficio", "garantía", "inclusión", "servicio"],
    "capital": ["capital asegurado", "monto asegurado", "suma asegurada", "valor asegurado", "límite", "importe", "cantidad", "capital", "monto", "suma", "valor"],
    "prima": ["prima", "precio", "costo", "tarifa", "valor", "cuota", "monto", "importe", "pago"],
    "deducible": ["deducible", "franquicia", "copago", "participación", "exceso", "coaseguro"],
    "reembolso": ["reembolso", "devolución", "pago", "compensación", "restitución", "reintegro"],
    "hospitalización": ["hospitalización", "internación", "ingreso", "estancia hospitalaria", "hospital"],
    "ambulatorio": ["ambulatorio", "consulta externa", "tratamiento ambulatorio", "externo"],
    "emergencia": ["emergencia", "urgencia", "accidente", "siniestro", "crisis", "urgente"],
    "exclusión": ["exclusión", "no cubre", "excluye", "no incluye", "excepto", "limitación", "restricción"],
    "covid": ["covid", "coronavirus", "pandemia", "enfermedad epidémica", "covid-19", "sars-cov-2"],
    "maternidad": ["maternidad", "embarazo", "parto", "nacimiento", "control prenatal", "gestación"],
    "parto": ["parto natural", "parto normal", "parto", "nacimiento", "cesárea", "parto vaginal"],
    "precio": ["precio", "costo", "tarifa", "valor", "monto", "importe", "cuota", "pago", "USD", "Bs", "bolivianos", "dólares"],
    "total": ["total", "suma total", "monto total", "importe total", "valor total", "gran total"],
    "anual": ["anual", "por año", "anualmente", "año", "anuales"],
    "mensual": ["mensual", "por mes", "mensualmente", "mes", "mensuales"],
}

def detectar_tipo_consulta(pregunta):
    pregunta_lower = pregunta.lower()
    tipos_detectados = []

    for termino, sinonimos in TERMINOS_SEGUROS.items():
        if any(s in pregunta_lower for s in [termino] + sinonimos[:2]):
            tipos_detectados.append(termino)

    if "modalidad" in tipos_detectados:
        return "MODALIDAD"
    elif "geográfico" in tipos_detectados or "nacional" in tipos_detectados or "internacional" in tipos_detectados:
        return "COBERTURA_GEOGRÁFICA"
    elif "capital" in tipos_detectados or "monto" in tipos_detectados or "precio" in tipos_detectados or "total" in tipos_detectados:
        return "MONTO_CAPITAL"
    elif "reembolso" in tipos_detectados:
        return "REEMBOLSO"
    elif "cobertura" in tipos_detectados:
        return "COBERTURA_GENERAL"

    return "GENERAL"

def preguntar_llm_optimizada(pregunta, contexto, client):
    tipo_consulta = detectar_tipo_consulta(pregunta)

    sinonimos_relevantes = []
    for termino, sinonimos in TERMINOS_SEGUROS.items():
        if termino in pregunta.lower() or any(s in pregunta.lower() for s in sinonimos[:4]):
            sinonimos_relevantes.extend(sinonimos[:4])

    sinonimos_texto = ", ".join(set(sinonimos_relevantes[:8]))

    prompt = f"""
Eres experto en pólizas de seguros. Analiza la siguiente consulta y busca información que coincida o similitudes.

DOCUMENTO DE LA PÓLIZA:
{contexto}

CONSULTA: "{pregunta}"

TIPO DE CONSULTA DETECTADA: {tipo_consulta}
TÉRMINOS RELACIONADOS PARA BÚSQUEDA: {sinonimos_texto}

INSTRUCCIONES ESPECÍFICAS PARA {tipo_consulta}:

{get_instrucciones_especificas(tipo_consulta)}

REGLAS GENERALES:
1. Busca por SIGNIFICADO, no solo por palabras exactas
2. Usa los términos relacionados para búsqueda ampliada
3. Copia EXACTAMENTE el texto relevante del documento, incluyendo NÚMEROS COMPLETOS con decimales
4. NO inventes, NO interpretes, NO resumas
5. Si no encuentras, responde EXACTAMENTE: "no encontrado"
6. NO agregues "Respuesta:", ni explicaciones, ni formato
7. IMPORTANTE: Si hay montos, precios o cifras, incluye TODOS los dígitos, incluyendo decimales y símbolos de moneda

EJEMPLOS CORRECTOS:
- MONTO_CAPITAL: "USD 5,000.00 según detalle" o "Bs. 35,000.00"
- PRECIO: "Prima anual: USD 1,250.50" o "Monto total: Bs. 8,750.00"

RESPUESTA DIRECTA DEL DOCUMENTO (INCLUYE NÚMEROS COMPLETOS):
"""

    for intento in range(MAX_REINTENTOS + 1):
        try:
            response = client.chat.complete(
                model=LLM_MODEL,
                messages=[{"role": "user", "content": prompt}],
                temperature=0.15,
                max_tokens=150
            )
            respuesta = response.choices[0].message.content.strip()
            return clean_response_completo(respuesta)
        except Exception as e:
            if intento < MAX_REINTENTOS and ("rate_limited" in str(e) or "429" in str(e)):
                espera = (2 ** intento) * 0.3 + random.uniform(0.3, 0.8)
                time.sleep(espera)
            else:
                return "error"

    return "error: rate limit"

def get_instrucciones_especificas(tipo_consulta):
    instrucciones = {
        "MODALIDAD": """
        • Busca términos: "modalidad", "tipo de póliza", "sistema", "plan"
        • Especifica si es: mixto, abierto, cerrado, combinado, dual
        • Incluye detalles del sistema de atención
        """,
        "COBERTURA_GEOGRÁFICA": """
        • Busca términos: "cobertura geográfica", "ámbito", "departamentos", "nacional", "internacional"
        • Especifica zonas, regiones, países cubiertos
        • Menciona límites territoriales si existen
        """,
        "MONTO_CAPITAL": """
        • Busca términos: "capital", "monto", "suma", "valor", "USD", "Bs.", "límite", "precio", "prima", "total"
        • Extrae números exactos con su moneda COMPLETOS (incluyendo decimales)
        • Incluye condiciones si las hay
        • Menciona si es anual, mensual o por evento
        • IMPORTANTE: Incluye TODOS los dígitos y símbolos de moneda
        """,
        "REEMBOLSO": """
        • Busca términos: "reembolso", "porcentaje", "%", "cobertura", "pago"
        • Extrae porcentajes exactos y condiciones
        • Especifica plazos y modalidades
        """,
        "COBERTURA_GENERAL": """
        • Busca términos específicos de la consulta
        • Extrae condiciones, límites, inclusiones
        • Especifica detalles relevantes
        """,
        "GENERAL": """
        • Busca información relacionada con la consulta
        • Usa términos equivalentes del sector
        • Extrae información precisa y relevante
        """
    }
    return instrucciones.get(tipo_consulta, instrucciones["GENERAL"])

def clean_response_completo(respuesta):
    """Limpia respuesta SIN cortar números/montos"""
    if not respuesta:
        return "no encontrado"

    respuesta_lower = respuesta.lower()

    no_encontrado_patterns = [
        "no encontrado", "no se encuentra", "no existe", "no hay información",
        "no aparece", "no figura", "no se menciona", "no se halla"
    ]

    for pattern in no_encontrado_patterns:
        if pattern in respuesta_lower:
            return "no encontrado"

    # Eliminar prefijos pero conservar números
    prefixes_to_remove = [
        r'^respuesta[:\s]*',
        r'^resultado[:\s]*',
        r'^encontrado[:\s]*',
        r'^información[:\s]*',
        r'^\s*[-*•]\s*',
        r'^\s*\d+[\.\)]\s*'
    ]

    for prefix in prefixes_to_remove:
        respuesta = re.sub(prefix, '', respuesta, flags=re.IGNORECASE)

    respuesta = re.sub(r'\*\*.*?\*\*', '', respuesta)
    respuesta = re.sub(r'__.*?__', '', respuesta)
    respuesta = re.sub(r'\s+', ' ', respuesta).strip()

    if len(respuesta) < 10 or len(respuesta.split()) < 3:
        return "no encontrado"

    # ✅ NO CORTAR si hay números/montos
    if len(respuesta) > 500:
        if re.search(r'[\d\.,]+|USD|Bs\.|Bolivianos|Dólares', respuesta):
            sentences = re.split(r'[.!?]', respuesta)
            for sentence in sentences:
                if re.search(r'[\d\.,]+', sentence) and len(sentence) > 20:
                    return sentence.strip() + "."

    return respuesta[:500] + "..." if len(respuesta) > 500 else respuesta

# =========================
# FUNCIONES DE PREPARACIÓN OPTIMIZADAS
# =========================
def preparar_documento(pdf_path, client, task_id=None, stage_prefix=""):
    if task_id:
        update_task(task_id, status="procesando_pdf", message=f"{stage_prefix}Extrayendo texto...")

    texto = extraer_texto_pdf(pdf_path, task_id, stage_prefix)

    if task_id:
        update_task(task_id, message=f"{stage_prefix}Dividiendo en fragmentos...")

    chunks = dividir_texto(texto, chunk_size=CHUNK_SIZE, overlap=OVERLAP)

    if task_id:
        update_task(task_id, message=f"{stage_prefix}Generando embeddings...")

    embeddings_list = []

    for i in range(0, len(chunks), BATCH_SIZE_EMBEDDINGS):
        if task_id:
            progress_base = 20 if "Póliza 1" in stage_prefix else 45
            progress = progress_base + (i / len(chunks)) * 20
            update_task(task_id, progress=progress, message=f"{stage_prefix}Embeddings {i}/{len(chunks)}")

        batch = chunks[i:i+BATCH_SIZE_EMBEDDINGS]
        try:
            batch_embeddings = generar_embeddings_con_reintentos(batch, client)
            embeddings_list.append(batch_embeddings)
            time.sleep(0.2)
        except Exception:
            continue

    embeddings = np.vstack(embeddings_list) if embeddings_list else np.array([])
    return chunks, embeddings

def procesar_consulta_dual(fila, consulta, chunks1, embeddings1, chunks2, embeddings2, chunks_ocr, embeddings_ocr, client, usar_ocr):
    resultado = {'fila': fila, 'consulta': consulta}

    # Procesar con PDF 1
    try:
        contexto_chunks1 = buscar_contexto(consulta, chunks1, embeddings1, embeddings_cache_pdf1, client, top_k=TOP_K, umbral=UMBRAL_SIMILITUD)
        if contexto_chunks1:
            contexto_texto1 = "\n\n".join(contexto_chunks1[:3])
            respuesta1 = preguntar_llm_optimizada(consulta, contexto_texto1, client)
        else:
            respuesta1 = "no encontrado"
    except Exception as e:
        respuesta1 = f"error: {str(e)[:50]}"

    # Procesar con PDF 2
    try:
        contexto_chunks2 = buscar_contexto(consulta, chunks2, embeddings2, embeddings_cache_pdf2, client, top_k=TOP_K, umbral=UMBRAL_SIMILITUD)
        if contexto_chunks2:
            contexto_texto2 = "\n\n".join(contexto_chunks2[:4])
            respuesta2 = preguntar_llm_optimizada(consulta, contexto_texto2, client)
        else:
            respuesta2 = "no encontrado"
    except Exception as e:
        respuesta2 = f"error: {str(e)[:50]}"

    # Procesar con OCR si está habilitado
    respuesta3 = None
    if usar_ocr and chunks_ocr is not None and len(embeddings_ocr) > 0:
        try:
            contexto_chunks3 = buscar_contexto(consulta, chunks_ocr, embeddings_ocr, embeddings_cache_ocr, client, top_k=TOP_K, umbral=UMBRAL_SIMILITUD)
            if contexto_chunks3:
                contexto_texto3 = "\n\n".join(contexto_chunks3[:4])
                respuesta3 = preguntar_llm_optimizada(consulta, contexto_texto3, client)
            else:
                respuesta3 = "no encontrado"
        except Exception as e:
            respuesta3 = f"error: {str(e)[:50]}"

    resultado['respuesta_pdf1'] = respuesta1
    resultado['respuesta_pdf2'] = respuesta2
    resultado['respuesta_ocr'] = respuesta3

    return resultado

# =========================
# PROCESAMIENTO EN BACKGROUND (MODIFICADO PARA DUPLICAR HOJA)
# =========================
def procesar_documentos_task(task_id, pdf1_path, pdf2_path, excel_path, pdf3_path=None):
    """Procesamiento en background con actualización de progreso"""
    try:
        # Inicializar cliente Mistral
        client = inicializar_mistral()
        update_task(task_id, progress=3, message="✅ Conexión con Mistral AI establecida")
        time.sleep(0.5)

        # Cargar Excel + ✅ DUPLICAR HOJA INMEDIATAMENTE
        update_task(task_id, status="cargando_excel", progress=5, message="Leyendo archivo Excel...")
        wb = load_workbook(excel_path)
        ws = wb.active
        # ✅ CREAR COPIA EXACTA DE LA HOJA ORIGINAL AL INICIO
        ws_cumple = duplicar_hoja(wb, ws, f"{ws.title}_Cumple")
        vector = construir_vector(ws)
        total_filas = len(vector)

        if total_filas == 0:
            raise Exception("No se encontraron filas para procesar en el Excel")

        update_task(task_id, progress=8, message=f"✅ Excel cargado: {total_filas} preguntas")
        time.sleep(0.5)

        # Preparar PDF 1
        update_task(task_id, progress=10, message="📄 Procesando Póliza 1...")
        chunks1, embeddings1 = preparar_documento(pdf1_path, client, task_id, "Póliza 1: ")

        # Preparar PDF 2
        update_task(task_id, progress=40, message="📄 Procesando Póliza 2...")
        chunks2, embeddings2 = preparar_documento(pdf2_path, client, task_id, "Póliza 2: ")

        # Preparar PDF OCR si existe
        usar_ocr = pdf3_path is not None
        chunks_ocr, embeddings_ocr = None, np.array([])
        if usar_ocr:
            update_task(task_id, progress=65, message="🖼️ Aplicando OCR al PDF escaneado...")
            chunks_ocr, embeddings_ocr = preparar_documento(pdf3_path, client, task_id, "OCR: ")

        # Procesar filas
        update_task(task_id, status="procesando_preguntas", progress=75, message=f"🧠 Procesando {total_filas} preguntas con IA...")
        resultados = []
        inicio_proc = time.time()

        for i, (fila, texto_consulta) in enumerate(vector, 1):
            # Actualizar progreso
            progress = 75 + (i / total_filas) * 20
            update_task(task_id, progress=progress, message=f"Pregunta {i}/{total_filas}: {texto_consulta[:40]}...")

            resultado = procesar_consulta_dual(
                fila, texto_consulta,
                chunks1, embeddings1,
                chunks2, embeddings2,
                chunks_ocr, embeddings_ocr,
                client,
                usar_ocr
            )
            resultados.append(resultado)

            if i < total_filas:
                time.sleep(PAUSA_ENTRE_CONSULTAS)
            time.sleep(0.15)

        tiempo_proc = time.time() - inicio_proc

        # Guardar resultados en AMBAS HOJAS
        update_task(task_id, status="guardando_resultados", progress=97, message="💾 Guardando resultados en Excel...")

        # Hoja original: respuestas detalladas
        ws.cell(row=FILA_INICIO-1, column=COL_F, value="DOCUMENTO")
        ws.cell(row=FILA_INICIO-1, column=COL_G, value=os.path.basename(pdf1_path).replace(".pdf", ""))
        ws.cell(row=FILA_INICIO-1, column=COL_H, value=os.path.basename(pdf2_path).replace(".pdf", ""))
        if usar_ocr:
            ws.cell(row=FILA_INICIO-1, column=COL_I, value="OCR: " + os.path.basename(pdf3_path).replace(".pdf", ""))

        # Hoja duplicada: encabezados "Cumple/No Cumple"
        ws_cumple.cell(row=FILA_INICIO-1, column=COL_F, value="DOCUMENTO")
        ws_cumple.cell(row=FILA_INICIO-1, column=COL_G, value=f"Cumple: {os.path.basename(pdf1_path).replace('.pdf', '')}")
        ws_cumple.cell(row=FILA_INICIO-1, column=COL_H, value=f"Cumple: {os.path.basename(pdf2_path).replace('.pdf', '')}")
        if usar_ocr:
            ws_cumple.cell(row=FILA_INICIO-1, column=COL_I, value=f"Cumple: OCR {os.path.basename(pdf3_path).replace('.pdf', '')}")

        # Llenar ambas hojas
        for r in resultados:
            # Hoja original: respuestas detalladas
            escribir_en_celda(ws, r['fila'], COL_G, r['respuesta_pdf1'])
            escribir_en_celda(ws, r['fila'], COL_H, r['respuesta_pdf2'])
            if usar_ocr and r.get('respuesta_ocr'):
                escribir_en_celda(ws, r['fila'], COL_I, r['respuesta_ocr'])

            # Hoja duplicada: Cumple/No Cumple
            # PDF 1
            if "no encontrado" in str(r['respuesta_pdf1']).lower() or "error" in str(r['respuesta_pdf1']).lower():
                escribir_en_celda(ws_cumple, r['fila'], COL_G, "No Cumple")
            else:
                escribir_en_celda(ws_cumple, r['fila'], COL_G, "Cumple")

            # PDF 2
            if "no encontrado" in str(r['respuesta_pdf2']).lower() or "error" in str(r['respuesta_pdf2']).lower():
                escribir_en_celda(ws_cumple, r['fila'], COL_H, "No Cumple")
            else:
                escribir_en_celda(ws_cumple, r['fila'], COL_H, "Cumple")

            # OCR
            if usar_ocr and r.get('respuesta_ocr'):
                if "no encontrado" in str(r['respuesta_ocr']).lower() or "error" in str(r['respuesta_ocr']).lower():
                    escribir_en_celda(ws_cumple, r['fila'], COL_I, "No Cumple")
                else:
                    escribir_en_celda(ws_cumple, r['fila'], COL_I, "Cumple")

        # Guardar archivo
        output_filename = f"resultados_{task_id}.xlsx"
        output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
        wb.save(output_path)

        # Calcular estadísticas
        contadores_pdf1 = {'encontrado': 0, 'no_encontrado': 0, 'error': 0}
        contadores_pdf2 = {'encontrado': 0, 'no_encontrado': 0, 'error': 0}
        contadores_ocr = {'encontrado': 0, 'no_encontrado': 0, 'error': 0} if usar_ocr else None

        for r in resultados:
            if "no encontrado" in str(r['respuesta_pdf1']).lower():
                contadores_pdf1['no_encontrado'] += 1
            elif "error" in str(r['respuesta_pdf1']).lower():
                contadores_pdf1['error'] += 1
            else:
                contadores_pdf1['encontrado'] += 1

            if "no encontrado" in str(r['respuesta_pdf2']).lower():
                contadores_pdf2['no_encontrado'] += 1
            elif "error" in str(r['respuesta_pdf2']).lower():
                contadores_pdf2['error'] += 1
            else:
                contadores_pdf2['encontrado'] += 1

            if usar_ocr and r.get('respuesta_ocr'):
                if "no encontrado" in str(r['respuesta_ocr']).lower():
                    contadores_ocr['no_encontrado'] += 1
                elif "error" in str(r['respuesta_ocr']).lower():
                    contadores_ocr['error'] += 1
                else:
                    contadores_ocr['encontrado'] += 1

        # Limpiar caches
        embeddings_cache_pdf1.clear()
        embeddings_cache_pdf2.clear()
        embeddings_cache_ocr.clear()

        # Resultado final
        update_task(task_id, status="completado", progress=100, message="✅ ¡Procesamiento completado!")
        app.config['TASKS'][task_id]['result'] = {
            'tiempo_min': tiempo_proc / 60,
            'velocidad': (total_filas / tiempo_proc) * 60 if tiempo_proc > 0 else 0,
            'total_filas': total_filas,
            'pdf1_name': os.path.basename(pdf1_path),
            'pdf2_name': os.path.basename(pdf2_path),
            'pdf3_name': os.path.basename(pdf3_path) if usar_ocr else None,
            'contadores_pdf1': contadores_pdf1,
            'contadores_pdf2': contadores_pdf2,
            'contadores_ocr': contadores_ocr,
            'resultados': resultados[:10],
            'output_filename': output_filename
        }

    except Exception as e:
        update_task(task_id, status="error", progress=0, message=f"❌ Error: {str(e)[:100]}")
        app.config['TASKS'][task_id]['error'] = str(e)

# =========================
# RUTAS FLASK
# =========================

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/procesar', methods=['POST'])
def procesar():
    try:
        if 'pdf1' not in request.files or 'pdf2' not in request.files or 'excel' not in request.files:
            flash('❌ Por favor, sube todos los archivos requeridos.', 'error')
            return redirect(url_for('index'))

        pdf1 = request.files['pdf1']
        pdf2 = request.files['pdf2']
        excel = request.files['excel']
        pdf3 = request.files.get('pdf3')

        if pdf1.filename == '' or pdf2.filename == '' or excel.filename == '':
            flash('❌ Por favor, selecciona todos los archivos requeridos.', 'error')
            return redirect(url_for('index'))

        # Crear directorio temporal
        tmpdir = tempfile.mkdtemp()

        # Guardar archivos
        pdf1_path = os.path.join(tmpdir, secure_filename(pdf1.filename))
        pdf2_path = os.path.join(tmpdir, secure_filename(pdf2.filename))
        excel_path = os.path.join(tmpdir, secure_filename(excel.filename))

        pdf1.save(pdf1_path)
        pdf2.save(pdf2_path)
        excel.save(excel_path)

        pdf3_path = None
        if pdf3 and pdf3.filename != '':
            pdf3_path = os.path.join(tmpdir, secure_filename(pdf3.filename))
            pdf3.save(pdf3_path)

        # Crear ID de tarea único
        task_id = f"task_{int(time.time())}_{os.urandom(4).hex()}"
        app.config['TASKS'][task_id] = {
            'status': 'iniciando',
            'progress': 0,
            'message': 'Iniciando procesamiento...',
            'created_at': datetime.now().isoformat()
        }

        # Iniciar procesamiento en background
        thread = threading.Thread(
            target=procesar_documentos_task,
            args=(task_id, pdf1_path, pdf2_path, excel_path, pdf3_path),
            daemon=True
        )
        thread.start()

        return redirect(url_for('processing', task_id=task_id))

    except Exception as e:
        flash(f'❌ Error al iniciar procesamiento: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/processing/<task_id>')
def processing(task_id):
    if task_id not in app.config['TASKS']:
        flash('❌ Tarea no encontrada', 'error')
        return redirect(url_for('index'))

    return render_template('processing.html', task_id=task_id)

@app.route('/status/<task_id>')
def status(task_id):
    task = app.config['TASKS'].get(task_id)
    if not task:
        return jsonify({'status': 'not_found', 'progress': 0, 'message': 'Tarea no encontrada'})

    # Limpiar tareas antiguas (> 1 hora)
    created_at = datetime.fromisoformat(task['created_at'])
    if (datetime.now() - created_at).total_seconds() > 3600:
        app.config['TASKS'].pop(task_id, None)
        return jsonify({'status': 'expired', 'progress': 0, 'message': 'Tarea expirada'})

    return jsonify(task)

@app.route('/resultados/<task_id>')
def resultados(task_id):
    task = app.config['TASKS'].get(task_id)
    if not task or task.get('status') != 'completado':
        flash('❌ Resultados no disponibles aún', 'error')
        return redirect(url_for('index'))

    return render_template('results.html', **task['result'])

@app.route('/descargar/<filename>')
def descargar(filename):
    try:
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        if os.path.exists(file_path):
            return send_file(file_path, as_attachment=True)
        else:
            flash('❌ Archivo no encontrado.', 'error')
            return redirect(url_for('index'))
    except Exception as e:
        flash(f'❌ Error al descargar: {str(e)}', 'error')
        return redirect(url_for('index'))

# =========================
# EJECUCIÓN
# =========================
if __name__ == '__main__':
    # Crear carpetas necesarias
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

    # Cargar variables de entorno
    from dotenv import load_dotenv
    load_dotenv()

    app.run(host='0.0.0.0', port=8501, debug=True)
